import pandas as pd
import numpy as np
import openpyxl


class Ans:
    def __init__(self, path):
        self.path = path
        print(self.path)

    def readFile(self):
        list_md = self.read_header()
        data = pd.ExcelFile(self.path).parse(0)
        list_question = list(data.iloc[:, 0])
        self.ans = dict()
        # self.ans['123'] = list(data.iloc[:, 1])
        # self.ans['393'] = list(data.iloc[:, 2])
        for i in range(len(list_md)):
            self.ans[str(list_md[i])] = list(data.iloc[:, i + 1])
        return self.ans

    def readAns(self, data):
        self.readFile()
        try:
            return self.ans[data]
        except:
            return []

    def read_header(self):
        wb_obj = openpyxl.load_workbook(self.path)
        sheet_obj = wb_obj.active
        list_md = []
        col = 2
        while sheet_obj.cell(row=1, column=col).value is not None:
            list_md.append(sheet_obj.cell(row=1, column=col).value)
            col += 1
        return list_md

    def write_data(self, mssv, value):
        data = pd.ExcelFile(self.path).parse(0)
        students = list(data.iloc[:, 0])
        index = students.index(int(mssv)) + 2
        cel = f'B{index}'
        wb_obj = openpyxl.load_workbook(self.path)
        sheet_obj = wb_obj.active
        sheet_obj[cel].value = value
        wb_obj.save(self.path)
